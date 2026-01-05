import olefile
import xlrd
from openpyxl import Workbook
import os
import re
 
# Define the folder path
folder_path = 'tournament-files/2024 National MS'
print(os.getcwd())
 
# List all files in the folder
files = os.listdir(folder_path)
 
# Define a pattern to match the filenames based on the provided structure
pattern = re.compile(r'.*\.(S\d+[A-Z])$')

def extract_sheets(file_path, output_path):
    # Open the OLE2 file
    ole = olefile.OleFileIO(file_path)

    # Check if the Workbook stream is present
    if ole.exists('Workbook'):
        # Extract the Workbook stream
        stream = ole.openstream('Workbook')
        
        # Save the stream content to a temporary .xls file
        with open('temp.xls', 'wb') as f:
            f.write(stream.read())

        # Open the temporary .xls file using xlrd
        workbook = xlrd.open_workbook('temp.xls')
        
        # Create a new Workbook using openpyxl
        new_workbook = Workbook()
        
        # Loop through the first three sheets (or as many as are available)
        for i in range(min(3, workbook.nsheets)):
            sheet = workbook.sheet_by_index(i)
            new_sheet = new_workbook.create_sheet(title=sheet.name)
            
            # Copy the content from the old sheet to the new sheet
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    new_sheet.cell(row=row+1, column=col+1, value=sheet.cell_value(row, col))
        
        # Remove the default empty sheet created by openpyxl
        if 'Sheet' in new_workbook.sheetnames:
            del new_workbook['Sheet']
        
        # Save the new workbook to an .xlsx file
        new_workbook.save(output_path)

        print("Sheets extracted successfully to {}.".format(output_path))

    else:
        print("Workbook stream not found for OLE2 file {}.".format(file_path))

    # Clean up
    ole.close()

# Loop through all files in the folder
conversion_results = []
for file in files:
    file_path = os.path.join(folder_path, file)
    if pattern.match(file):
        try:
            output_path = file_path + '.xlsx'
            extract_sheets(file_path, output_path)
            conversion_results.append(f"Extracted sheets from {file} to {output_path}")
        except Exception as e:
            conversion_results.append(f"Error processing {file}: {str(e)}")
    else:
        conversion_results.append(f"Skipping non-Excel file: {file_path}")
 
# Print conversion results
for result in conversion_results:
    print(result)